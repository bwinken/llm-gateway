"""Tests for the admin usage-export xlsx endpoint and analytics aggregations."""

from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.schema import AppOwner, UsageLog, User
from app.services.analytics import (
    build_monthly_report,
    iter_months,
    parse_ym,
    top_users_with_delta,
)
from tests.conftest import web_auth_header


def _dt(y: int, m: int, d: int, h: int = 12) -> datetime:
    return datetime(y, m, d, h, 0, 0, tzinfo=timezone.utc)


@pytest.fixture()
def seeded(db_session):
    """Seed users (humans + apps) and usage logs spanning Jan–Mar 2026."""
    users = [
        User(username="alice", display_name="Alice A.", org_code="SALES",
             api_key="sk-a", is_admin=False),
        User(username="bob", display_name="Bob B.", org_code="ENG",
             api_key="sk-b", is_admin=False),
        User(username="carol", display_name="Carol C.", org_code="ENG",
             api_key="sk-c", is_admin=False),
        User(username="app_backend", display_name="", org_code="",
             api_key="sk-app1", is_admin=False),
    ]
    for u in users:
        db_session.add(u)
    db_session.commit()
    for u in users:
        db_session.refresh(u)
    alice, bob, carol, app_backend = users

    # Ownership: bob owns app_backend
    db_session.add(AppOwner(app_id=app_backend.id, owner_id=bob.id))
    db_session.commit()

    logs = [
        # January — alice leads, bob mid, carol low. Alice's spend is split
        # across all three backends so the per-backend pivot has real data.
        UsageLog(user_id=alice.id, input_tokens=100, output_tokens=50,
                 cost_usd=Decimal("1.00"), created_at=_dt(2026, 1, 5)),
        UsageLog(user_id=alice.id, input_tokens=200, output_tokens=100,
                 cost_usd=Decimal("2.00"), created_at=_dt(2026, 1, 12),
                 backend="azure", endpoint="/azure/v1/chat/completions"),
        UsageLog(user_id=bob.id, input_tokens=150, output_tokens=75,
                 cost_usd=Decimal("1.50"), created_at=_dt(2026, 1, 8),
                 backend="bedrock", endpoint="/aws/v1/chat/completions"),
        UsageLog(user_id=carol.id, input_tokens=50, output_tokens=25,
                 cost_usd=Decimal("0.50"), created_at=_dt(2026, 1, 20)),
        UsageLog(user_id=app_backend.id, input_tokens=500, output_tokens=250,
                 cost_usd=Decimal("5.00"), created_at=_dt(2026, 1, 15)),
        # February — bob overtakes alice
        UsageLog(user_id=bob.id, input_tokens=400, output_tokens=200,
                 cost_usd=Decimal("4.00"), created_at=_dt(2026, 2, 3)),
        UsageLog(user_id=bob.id, input_tokens=200, output_tokens=100,
                 cost_usd=Decimal("2.00"), created_at=_dt(2026, 2, 18)),
        UsageLog(user_id=alice.id, input_tokens=100, output_tokens=50,
                 cost_usd=Decimal("1.00"), created_at=_dt(2026, 2, 10)),
        # March — only carol active (new month, no prev-top-10 for her)
        UsageLog(user_id=carol.id, input_tokens=300, output_tokens=150,
                 cost_usd=Decimal("3.00"), created_at=_dt(2026, 3, 4)),
    ]
    for log in logs:
        db_session.add(log)
    db_session.commit()
    return {
        "alice": alice, "bob": bob, "carol": carol, "app_backend": app_backend,
    }


# ──────────────────────────────── Unit tests ─────────────────────────────────

class TestParsing:
    def test_parse_valid(self):
        assert parse_ym("2026-04") == (2026, 4)

    def test_parse_invalid(self):
        with pytest.raises(ValueError):
            parse_ym("2026-13")
        with pytest.raises(ValueError):
            parse_ym("2026")

    def test_iter_months_inclusive(self):
        assert iter_months("2026-01", "2026-03") == [(2026, 1), (2026, 2), (2026, 3)]

    def test_iter_months_cross_year(self):
        assert iter_months("2025-11", "2026-02") == [
            (2025, 11), (2025, 12), (2026, 1), (2026, 2),
        ]

    def test_iter_months_rejects_reversed(self):
        with pytest.raises(ValueError):
            iter_months("2026-03", "2026-01")


class TestMonthlyReport:
    def test_summary_counts_humans_only_for_dau_mau(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-01")
        s = report.breakdowns[0].summary
        # Alice + Bob + Carol = 3 humans; app_backend excluded from MAU
        assert s["mau"] == 3
        # Totals include app traffic: 5 log rows in Jan
        assert s["requests"] == 5
        # Peak DAU is 1 per day (each user logged once on a distinct day)
        assert s["dau_peak"] == 1

    def test_by_department_excludes_apps(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-01")
        depts = {d["department"]: d for d in report.breakdowns[0].by_department}
        assert set(depts.keys()) == {"SALES", "ENG"}
        # ENG = bob + carol = 2 users
        assert depts["ENG"]["users"] == 2
        assert depts["SALES"]["users"] == 1

    def test_by_app_lists_only_app_prefixed_users(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-01")
        apps = report.breakdowns[0].by_app
        assert len(apps) == 1
        assert apps[0]["app"] == "app_backend"
        assert apps[0]["owners"] == ["bob"]
        assert apps[0]["cost_usd"] == 5.0

    def test_user_ranking_excludes_apps(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-01")
        ranking = report.breakdowns[0].user_ranking
        names = [r["username"] for r in ranking]
        assert "app_backend" not in names
        # Alice should rank 1 in January (3.00 cost)
        assert ranking[0]["username"] == "alice"
        assert ranking[0]["rank"] == 1


class TestByUserBackend:
    """Per-user cost split across vLLM / Azure / Bedrock (the export the
    chargeback flow consumes)."""

    def test_backend_costs_pivoted_per_user(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-01")
        rows = {r["username"]: r for r in report.breakdowns[0].by_user_backend}

        alice = rows["alice"]
        assert alice["vllm_cost_usd"] == 1.0     # the untagged (default) row
        assert alice["azure_cost_usd"] == 2.0
        assert alice["bedrock_cost_usd"] == 0.0
        assert alice["total_cost_usd"] == 3.0
        assert alice["requests"] == 2

        bob = rows["bob"]
        assert bob["bedrock_cost_usd"] == 1.5
        assert bob["vllm_cost_usd"] == 0.0
        assert bob["total_cost_usd"] == 1.5

    def test_includes_app_accounts_with_flag(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-01")
        rows = {r["username"]: r for r in report.breakdowns[0].by_user_backend}
        assert rows["app_backend"]["is_app"] is True
        assert rows["app_backend"]["vllm_cost_usd"] == 5.0
        assert rows["alice"]["is_app"] is False

    def test_sorted_by_total_cost_desc(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-01")
        totals = [r["total_cost_usd"] for r in report.breakdowns[0].by_user_backend]
        assert totals == sorted(totals, reverse=True)

    def test_columns_sum_to_total(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-01")
        for r in report.breakdowns[0].by_user_backend:
            assert r["total_cost_usd"] == pytest.approx(
                r["vllm_cost_usd"] + r["azure_cost_usd"] + r["bedrock_cost_usd"]
            )

    def test_month_without_usage_is_empty(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-04", "2026-04")
        assert report.breakdowns[0].by_user_backend == []


class TestTopUsersDelta:
    def test_rank_movement_between_months(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-01", "2026-02")
        rows = top_users_with_delta(report, limit=10)

        jan = [r for r in rows if r["month"] == "2026-01"]
        feb = [r for r in rows if r["month"] == "2026-02"]

        # January: everyone is a new entrant (no prior month in the range)
        for r in jan:
            assert r["prev_rank"] is None
            assert r["rank_delta"] is None

        # February: alice was #1 in Jan, dropped to #2 in Feb
        # bob was #2 in Jan, climbed to #1 in Feb
        by_user = {r["username"]: r for r in feb}
        assert by_user["bob"]["rank"] == 1
        assert by_user["bob"]["prev_rank"] == 2
        assert by_user["bob"]["rank_delta"] == 1     # climbed one spot
        assert by_user["alice"]["rank"] == 2
        assert by_user["alice"]["rank_delta"] == -1  # dropped one spot

    def test_new_entrant_after_absence(self, db_session, seeded):
        report = build_monthly_report(db_session, "2026-02", "2026-03")
        rows = top_users_with_delta(report, limit=10)
        mar = [r for r in rows if r["month"] == "2026-03"]
        # Carol had no February usage → appears in March as NEW
        carol = next(r for r in mar if r["username"] == "carol")
        assert carol["prev_rank"] is None
        assert carol["rank_delta"] is None


# ─────────────────────────────── Endpoint tests ──────────────────────────────

class TestExportEndpoint:
    def test_admin_can_download_xlsx(self, client, db_session, admin_user, seeded):
        resp = client.get(
            "/admin/api/export/usage.xlsx",
            params={"from": "2026-01", "to": "2026-03"},
            headers=web_auth_header(sub=admin_user.username, scopes=["admin"]),
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert "usage_2026-01_to_2026-03.xlsx" in resp.headers["content-disposition"]

        wb = load_workbook(BytesIO(resp.content))
        assert wb.sheetnames == [
            "Summary", "By Department", "By App", "Top 10 Users", "Cost by Backend",
        ]

        # Summary sheet: 3 data rows (Jan/Feb/Mar) starting at row 5
        summary = wb["Summary"]
        months = [summary.cell(row=r, column=1).value for r in (5, 6, 7)]
        assert months == ["2026-01", "2026-02", "2026-03"]

        # By Department sheet has header at row 1, data from row 2
        depts = wb["By Department"]
        header = [c.value for c in next(depts.iter_rows(min_row=1, max_row=1))]
        assert header[0] == "Month" and header[1] == "Department"

        # By App sheet should include app_backend in January
        apps = wb["By App"]
        app_rows = [
            [c.value for c in row]
            for row in apps.iter_rows(min_row=2, values_only=False)
        ]
        assert any(row[1] == "app_backend" for row in app_rows)

    def test_cost_by_backend_sheet_rows(self, client, db_session, admin_user, seeded):
        resp = client.get(
            "/admin/api/export/usage.xlsx",
            params={"from": "2026-01", "to": "2026-01"},
            headers=web_auth_header(sub=admin_user.username, scopes=["admin"]),
        )
        assert resp.status_code == 200
        ws = load_workbook(BytesIO(resp.content))["Cost by Backend"]
        # Header at row 3, data from row 4: [Month, Account, ..., On-prem, Azure, AWS, Total]
        rows = {
            r[1]: r for r in ws.iter_rows(min_row=4, values_only=True) if r[1]
        }
        alice = rows["alice"]
        assert alice[8] == 1.0    # On-prem
        assert alice[9] == 2.0    # Azure
        assert alice[10] == 0.0   # AWS
        assert alice[11] == 3.0   # Total
        assert rows["app_backend"][4] == "App"

    def test_non_admin_denied(self, client, seeded):
        resp = client.get(
            "/admin/api/export/usage.xlsx",
            params={"from": "2026-01", "to": "2026-03"},
            headers=web_auth_header(sub="testuser", scopes=["read"]),
        )
        assert resp.status_code == 403

    def test_invalid_month_returns_400(self, client, admin_user):
        resp = client.get(
            "/admin/api/export/usage.xlsx",
            params={"from": "2026-13", "to": "2026-03"},
            headers=web_auth_header(sub=admin_user.username, scopes=["admin"]),
        )
        assert resp.status_code == 400

    def test_reversed_range_returns_400(self, client, admin_user):
        resp = client.get(
            "/admin/api/export/usage.xlsx",
            params={"from": "2026-03", "to": "2026-01"},
            headers=web_auth_header(sub=admin_user.username, scopes=["admin"]),
        )
        assert resp.status_code == 400

    def test_range_too_wide_returns_400(self, client, admin_user):
        resp = client.get(
            "/admin/api/export/usage.xlsx",
            params={"from": "2020-01", "to": "2026-12"},
            headers=web_auth_header(sub=admin_user.username, scopes=["admin"]),
        )
        assert resp.status_code == 400


class TestUserBackendCostsCsv:
    """CSV counterpart of the "Cost by Backend" sheet — one row per
    (month × account) with separate per-backend cost columns."""

    def _get(self, client, admin_user, **params):
        return client.get(
            "/admin/api/export/user-backend-costs.csv",
            params=params,
            headers=web_auth_header(sub=admin_user.username, scopes=["admin"]),
        )

    def test_admin_can_download_csv(self, client, db_session, admin_user, seeded):
        resp = self._get(client, admin_user, **{"from": "2026-01", "to": "2026-01"})
        assert resp.status_code == 200
        assert resp.headers["content-type"].startswith("text/csv")
        assert "user_backend_costs_2026-01_to_2026-01.csv" in resp.headers["content-disposition"]

        import csv as csv_mod
        import io

        reader = csv_mod.DictReader(io.StringIO(resp.text.lstrip("﻿")))
        rows = {r["username"]: r for r in reader}

        alice = rows["alice"]
        assert alice["month"] == "2026-01"
        assert alice["type"] == "user"
        assert float(alice["vllm_cost_usd"]) == 1.0
        assert float(alice["azure_cost_usd"]) == 2.0
        assert float(alice["bedrock_cost_usd"]) == 0.0
        assert float(alice["total_cost_usd"]) == 3.0

        bob = rows["bob"]
        assert float(bob["bedrock_cost_usd"]) == 1.5

        app = rows["app_backend"]
        assert app["type"] == "app"
        assert float(app["vllm_cost_usd"]) == 5.0

    def test_multi_month_rows_carry_month_column(self, client, db_session, admin_user, seeded):
        resp = self._get(client, admin_user, **{"from": "2026-01", "to": "2026-03"})
        assert resp.status_code == 200

        import csv as csv_mod
        import io

        reader = csv_mod.DictReader(io.StringIO(resp.text.lstrip("﻿")))
        months = {r["month"] for r in reader}
        assert months == {"2026-01", "2026-02", "2026-03"}

    def test_non_admin_denied(self, client, seeded):
        resp = client.get(
            "/admin/api/export/user-backend-costs.csv",
            params={"from": "2026-01", "to": "2026-01"},
            headers=web_auth_header(sub="testuser", scopes=["read"]),
        )
        assert resp.status_code == 403

    def test_invalid_month_returns_400(self, client, admin_user):
        resp = self._get(client, admin_user, **{"from": "2026-13", "to": "2026-01"})
        assert resp.status_code == 400

    def test_range_too_wide_returns_400(self, client, admin_user):
        resp = self._get(client, admin_user, **{"from": "2020-01", "to": "2026-12"})
        assert resp.status_code == 400
