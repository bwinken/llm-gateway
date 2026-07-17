"""
Build an xlsx workbook from a MonthlyReport (see app.services.analytics).

Sheets:
  1. Summary          — one row per month (reqs, tokens, cost, DAU/MAU)
  2. By Department    — (month × org_code) rows
  3. By App           — (month × app account) rows, with owners
  4. Top 10 Users     — (month × rank 1–10) rows with previous-month rank delta
  5. Cost by Backend  — (month × account) rows with the cost split across
                        On-prem (vLLM) / Azure / AWS Bedrock, humans and apps
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.services.analytics import MonthlyReport, top_users_with_delta


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4B5563")
_META_FONT = Font(italic=True, color="6B7280")


def _write_header(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def _autosize(ws: Worksheet, *, min_width: int = 10, max_width: int = 40) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        longest = max(
            (len(str(c.value)) for c in col_cells if c.value is not None),
            default=0,
        )
        ws.column_dimensions[letter].width = max(min_width, min(max_width, longest + 2))


def _build_summary(ws: Worksheet, report: MonthlyReport) -> None:
    ws.title = "Summary"
    ws.cell(
        row=1, column=1,
        value=f"Usage Summary — {report.from_ym} to {report.to_ym}",
    ).font = Font(bold=True, size=13)
    ws.cell(
        row=2, column=1,
        value="DAU = daily distinct human users (averaged over days in month); MAU = distinct human users in month. Apps excluded.",
    ).font = _META_FONT

    headers = [
        "Month", "Requests", "Input Tokens", "Output Tokens", "Total Tokens",
        "Cost (USD)", "MAU", "DAU (avg)", "DAU (peak)", "Active Days",
    ]
    _write_header(ws, 4, headers)

    for i, bd in enumerate(report.breakdowns, start=5):
        s = bd.summary
        ws.cell(row=i, column=1, value=bd.month)
        ws.cell(row=i, column=2, value=s["requests"])
        ws.cell(row=i, column=3, value=s["input_tokens"])
        ws.cell(row=i, column=4, value=s["output_tokens"])
        ws.cell(row=i, column=5, value=s["input_tokens"] + s["output_tokens"])
        ws.cell(row=i, column=6, value=s["cost_usd"]).number_format = "$#,##0.0000"
        ws.cell(row=i, column=7, value=s["mau"])
        ws.cell(row=i, column=8, value=s["dau_avg"]).number_format = "0.00"
        ws.cell(row=i, column=9, value=s["dau_peak"])
        ws.cell(row=i, column=10, value=f"{s['active_days']} / {s['days_in_month']}")

    _autosize(ws)
    ws.freeze_panes = "A5"


def _build_by_department(ws: Worksheet, report: MonthlyReport) -> None:
    ws.title = "By Department"
    headers = [
        "Month", "Department", "Users", "Requests",
        "Input Tokens", "Output Tokens", "Cost (USD)",
    ]
    _write_header(ws, 1, headers)

    row = 2
    for bd in report.breakdowns:
        for d in bd.by_department:
            ws.cell(row=row, column=1, value=bd.month)
            ws.cell(row=row, column=2, value=d["department"])
            ws.cell(row=row, column=3, value=d["users"])
            ws.cell(row=row, column=4, value=d["requests"])
            ws.cell(row=row, column=5, value=d["input_tokens"])
            ws.cell(row=row, column=6, value=d["output_tokens"])
            ws.cell(row=row, column=7, value=d["cost_usd"]).number_format = "$#,##0.0000"
            row += 1

    _autosize(ws)
    ws.freeze_panes = "A2"


def _build_by_app(ws: Worksheet, report: MonthlyReport) -> None:
    ws.title = "By App"
    headers = [
        "Month", "App", "Department", "Owners", "Requests",
        "Input Tokens", "Output Tokens", "Cost (USD)",
    ]
    _write_header(ws, 1, headers)

    row = 2
    for bd in report.breakdowns:
        for a in bd.by_app:
            ws.cell(row=row, column=1, value=bd.month)
            ws.cell(row=row, column=2, value=a["app"])
            ws.cell(row=row, column=3, value=a["org_code"])
            ws.cell(row=row, column=4, value=", ".join(a["owners"]))
            ws.cell(row=row, column=5, value=a["requests"])
            ws.cell(row=row, column=6, value=a["input_tokens"])
            ws.cell(row=row, column=7, value=a["output_tokens"])
            ws.cell(row=row, column=8, value=a["cost_usd"]).number_format = "$#,##0.0000"
            row += 1

    _autosize(ws)
    ws.freeze_panes = "A2"


def _build_top_users(ws: Worksheet, report: MonthlyReport, limit: int = 10) -> None:
    ws.title = f"Top {limit} Users"
    ws.cell(
        row=1, column=1,
        value=(
            "Δ Rank: negative = moved up, positive = moved down, blank = new entrant. "
            "Apps excluded."
        ),
    ).font = _META_FONT

    headers = [
        "Month", "Rank", "Prev Rank", "Δ Rank",
        "User", "Display Name", "Department",
        "Requests", "Input Tokens", "Output Tokens", "Cost (USD)",
    ]
    _write_header(ws, 3, headers)

    rows = top_users_with_delta(report, limit=limit)
    for i, r in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=r["month"])
        ws.cell(row=i, column=2, value=r["rank"])
        ws.cell(row=i, column=3, value=r["prev_rank"] if r["prev_rank"] is not None else "")
        if r["rank_delta"] is None:
            ws.cell(row=i, column=4, value="NEW")
        else:
            ws.cell(row=i, column=4, value=r["rank_delta"])
        ws.cell(row=i, column=5, value=r["username"])
        ws.cell(row=i, column=6, value=r["display_name"])
        ws.cell(row=i, column=7, value=r["org_code"])
        ws.cell(row=i, column=8, value=r["requests"])
        ws.cell(row=i, column=9, value=r["input_tokens"])
        ws.cell(row=i, column=10, value=r["output_tokens"])
        ws.cell(row=i, column=11, value=r["cost_usd"]).number_format = "$#,##0.0000"

    _autosize(ws)
    ws.freeze_panes = "A4"


def _build_cost_by_backend(ws: Worksheet, report: MonthlyReport) -> None:
    ws.title = "Cost by Backend"
    ws.cell(
        row=1, column=1,
        value=(
            "Per-account cost split by serving backend. Includes both human "
            "users and app_* accounts (see Type). Total = On-prem + Azure + AWS."
        ),
    ).font = _META_FONT

    headers = [
        "Month", "Account", "Display Name", "Department", "Type",
        "Requests", "Input Tokens", "Output Tokens",
        "On-prem Cost (USD)", "Azure Cost (USD)", "AWS Cost (USD)",
        "Total Cost (USD)",
    ]
    _write_header(ws, 3, headers)

    row = 4
    for bd in report.breakdowns:
        for u in bd.by_user_backend:
            ws.cell(row=row, column=1, value=bd.month)
            ws.cell(row=row, column=2, value=u["username"])
            ws.cell(row=row, column=3, value=u["display_name"])
            ws.cell(row=row, column=4, value=u["org_code"])
            ws.cell(row=row, column=5, value="App" if u["is_app"] else "User")
            ws.cell(row=row, column=6, value=u["requests"])
            ws.cell(row=row, column=7, value=u["input_tokens"])
            ws.cell(row=row, column=8, value=u["output_tokens"])
            ws.cell(row=row, column=9, value=u["vllm_cost_usd"]).number_format = "$#,##0.0000"
            ws.cell(row=row, column=10, value=u["azure_cost_usd"]).number_format = "$#,##0.0000"
            ws.cell(row=row, column=11, value=u["bedrock_cost_usd"]).number_format = "$#,##0.0000"
            ws.cell(row=row, column=12, value=u["total_cost_usd"]).number_format = "$#,##0.0000"
            row += 1

    _autosize(ws)
    ws.freeze_panes = "A4"


def build_workbook(report: MonthlyReport) -> bytes:
    """Render the full 5-sheet workbook as xlsx bytes."""
    wb = Workbook()
    _build_summary(wb.active, report)
    _build_by_department(wb.create_sheet(), report)
    _build_by_app(wb.create_sheet(), report)
    _build_top_users(wb.create_sheet(), report)
    _build_cost_by_backend(wb.create_sheet(), report)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
